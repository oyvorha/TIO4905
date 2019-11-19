import pandas as pd
from openpyxl import load_workbook

book = load_workbook("Output/output.xlsx")
writer = pd.ExcelWriter("Output/output.xlsx", engine='openpyxl')
writer.book = book

df_keys = []


def save_output(model, time, fixed, dynamic, station_obj):
    key = "solvable_instance_" + str(len(fixed.stations)) + '_' + str(len(fixed.vehicles))

    df = pd.DataFrame(columns=['Init #stations', 'No. of stations', 'No. of vehicles', 'Objective Value',
                                                 'Vehicle capacity', 'Station capacity', 'Time Horizon',
                                                 'Demand scenario', 'Solution time', 'Gap', 'weight violation',
                                                 'weight deviation', 'weight reward', 'reward weight dev',
                                                 'reward weight time'])
    for v in range(len(fixed.vehicles)):
        df['Start vehicle' + str(v)] = ""

    new_row = {'Init #stations': fixed.initial_size, 'No. of stations': len(fixed.stations)-2,
               'No. of vehicles': len(fixed.vehicles), 'Objective Value': model.objVal,
               'Vehicle capacity': fixed.vehicle_cap[0], 'Station capacity': fixed.station_cap[1],
               'Time Horizon': fixed.time_horizon, 'Demand scenario': fixed.demand_scenario,
               'Solution time': time, 'Gap': model.mipgap, 'weight violation': fixed.w_violation,
               'weight deviation': fixed.w_dev_obj, 'weight reward': fixed.w_reward,
               'reward weight dev': fixed.w_dev_reward, 'reward weight time': fixed.w_driving_time}

    for v in range(len(fixed.vehicles)):
        new_row['Start vehicle' + str(v)] = station_obj[dynamic.start_stations[v]].address

    for var in model.getVars():
        new_row[var.varName] = var.x
    df = df.append(new_row, ignore_index=True)

    if key in df_keys:
        start_row = writer.sheets[key].max_row
        df.to_excel(writer, startrow=start_row, index=False, header=False, sheet_name=key)
        writer.save()
    else:
        df.to_excel(writer, index=False, sheet_name=key)
        writer.save()
        df_keys.append(key)
